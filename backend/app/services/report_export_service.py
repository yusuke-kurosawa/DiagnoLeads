"""
Report Export Service

Handles exporting reports to various formats (PDF, Excel, CSV).

**Dependencies**:
- openpyxl: For Excel export
- reportlab or WeasyPrint: For PDF export (choose based on requirements)
- pandas (optional): For advanced Excel formatting
"""

import csv
import io
from typing import Dict, List, Any
from datetime import datetime


class ReportExportService:
    """
    Service for exporting report data to various formats

    Supports: PDF, Excel (XLSX), CSV
    """

    def export_to_csv(
        self, report_name: str, data_points: List[Dict[str, Any]]
    ) -> bytes:
        """
        Export report to CSV format

        Args:
            report_name: Name of the report
            data_points: List of data points with labels and values

        Returns:
            CSV file as bytes
        """
        output = io.StringIO()
        writer = csv.writer(output)

        # Extract all metric names
        if not data_points:
            return b""

        metrics = set()
        for dp in data_points:
            metrics.update(dp.get("values", {}).keys())

        # Write header
        header = ["Label"] + sorted(list(metrics))
        writer.writerow(header)

        # Write data rows
        for dp in data_points:
            label = dp.get("label", "Unknown")
            values = dp.get("values", {})
            row = [label] + [values.get(metric, 0) for metric in sorted(metrics)]
            writer.writerow(row)

        # Convert to bytes
        csv_content = output.getvalue()
        return csv_content.encode("utf-8")

    def export_to_excel(
        self,
        report_name: str,
        data_points: List[Dict[str, Any]],
        summary: Dict[str, Any],
        config: Dict[str, Any],
    ) -> bytes:
        """
        Export report to Excel format (XLSX)

        Requires: openpyxl

        Args:
            report_name: Name of the report
            data_points: List of data points
            summary: Summary statistics
            config: Report configuration

        Returns:
            Excel file as bytes
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError(
                "openpyxl is required for Excel export. Install with: pip install openpyxl"
            )

        wb = Workbook()

        # Data sheet
        ws_data = wb.active
        ws_data.title = "Report Data"

        # Title row
        ws_data["A1"] = report_name
        ws_data["A1"].font = Font(size=16, bold=True)
        ws_data["A2"] = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')}"

        # Extract metrics
        if data_points:
            metrics_set = set()
            for dp in data_points:
                metrics_set.update(dp.get("values", {}).keys())
            metrics = sorted(list(metrics_set))

            # Header row
            row = 4
            ws_data[f"A{row}"] = "Label"
            ws_data[f"A{row}"].font = Font(bold=True)
            ws_data[f"A{row}"].fill = PatternFill(
                start_color="CCE5FF", end_color="CCE5FF", fill_type="solid"
            )

            for col, metric in enumerate(metrics, start=2):
                cell = ws_data[f"{get_column_letter(col)}{row}"]
                cell.value = metric.replace("_", " ").title()
                cell.font = Font(bold=True)
                cell.fill = PatternFill(
                    start_color="CCE5FF", end_color="CCE5FF", fill_type="solid"
                )

            # Data rows
            for dp in data_points:
                row += 1
                label = dp.get("label", "Unknown")
                values = dp.get("values", {})

                ws_data[f"A{row}"] = label

                for col, metric in enumerate(metrics, start=2):
                    cell = ws_data[f"{get_column_letter(col)}{row}"]
                    cell.value = values.get(metric, 0)
                    # Format numbers
                    if isinstance(cell.value, float):
                        cell.number_format = "0.00"

        # Summary sheet
        ws_summary = wb.create_sheet("Summary")
        ws_summary["A1"] = "Summary Statistics"
        ws_summary["A1"].font = Font(size=14, bold=True)

        row = 3
        for key, value in summary.items():
            ws_summary[f"A{row}"] = key.replace("_", " ").title()
            ws_summary[f"A{row}"].font = Font(bold=True)
            ws_summary[f"B{row}"] = value
            row += 1

        # Configuration sheet
        ws_config = wb.create_sheet("Configuration")
        ws_config["A1"] = "Report Configuration"
        ws_config["A1"].font = Font(size=14, bold=True)

        row = 3
        ws_config[f"A{row}"] = "Metrics"
        ws_config[f"B{row}"] = ", ".join(config.get("metrics", []))
        row += 1

        if "filters" in config:
            ws_config[f"A{row}"] = "Filters"
            ws_config[f"B{row}"] = str(config["filters"])
            row += 1

        if "group_by" in config and config["group_by"]:
            ws_config[f"A{row}"] = "Group By"
            ws_config[f"B{row}"] = config["group_by"]
            row += 1

        # Auto-adjust column widths
        for sheet in wb.worksheets:
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return output.getvalue()

    def export_to_pdf(
        self,
        report_name: str,
        data_points: List[Dict[str, Any]],
        summary: Dict[str, Any],
        config: Dict[str, Any],
        include_charts: bool = True,
    ) -> bytes:
        """
        Export report to PDF format

        Requires: reportlab

        Args:
            report_name: Name of the report
            data_points: List of data points
            summary: Summary statistics
            config: Report configuration
            include_charts: Whether to include charts (requires matplotlib)

        Returns:
            PDF file as bytes
        """
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.platypus import (
                SimpleDocTemplate,
                Table,
                TableStyle,
                Paragraph,
                Spacer,
                PageBreak,
            )
            from reportlab.lib import colors
        except ImportError:
            raise ImportError(
                "reportlab is required for PDF export. Install with: pip install reportlab"
            )

        # Create PDF buffer
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        story = []
        styles = getSampleStyleSheet()

        # Title
        title_style = ParagraphStyle(
            "CustomTitle", parent=styles["Heading1"], fontSize=20, spaceAfter=30
        )
        story.append(Paragraph(report_name, title_style))

        # Generated timestamp
        story.append(
            Paragraph(
                f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')}",
                styles["Normal"],
            )
        )
        story.append(Spacer(1, 0.2 * inch))

        # Summary section
        story.append(Paragraph("Summary Statistics", styles["Heading2"]))
        summary_data = [["Metric", "Value"]]
        for key, value in summary.items():
            summary_data.append([key.replace("_", " ").title(), str(value)])

        summary_table = Table(summary_data, colWidths=[3 * inch, 2 * inch])
        summary_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 12),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ]
            )
        )
        story.append(summary_table)
        story.append(Spacer(1, 0.3 * inch))

        # Data section
        if data_points:
            story.append(Paragraph("Detailed Data", styles["Heading2"]))

            # Extract metrics
            metrics_set = set()
            for dp in data_points:
                metrics_set.update(dp.get("values", {}).keys())
            metrics = sorted(list(metrics_set))

            # Build table data
            table_data = [["Label"] + [m.replace("_", " ").title() for m in metrics]]

            for dp in data_points:
                label = dp.get("label", "Unknown")
                values = dp.get("values", {})
                row = [label] + [
                    f"{values.get(metric, 0):.2f}"
                    if isinstance(values.get(metric, 0), float)
                    else str(values.get(metric, 0))
                    for metric in metrics
                ]
                table_data.append(row)

            # Calculate column widths
            col_width = 5.5 * inch / (len(metrics) + 1)
            data_table = Table(table_data, colWidths=[col_width] * (len(metrics) + 1))
            data_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, 0), 10),
                        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                        ("GRID", (0, 0), (-1, -1), 1, colors.black),
                        (
                            "ROWBACKGROUNDS",
                            (0, 1),
                            (-1, -1),
                            [colors.white, colors.lightgrey],
                        ),
                    ]
                )
            )
            story.append(data_table)

        # Configuration section
        story.append(PageBreak())
        story.append(Paragraph("Report Configuration", styles["Heading2"]))

        config_text = f"""
        <b>Metrics:</b> {", ".join(config.get("metrics", []))}<br/>
        <b>Group By:</b> {config.get("group_by", "None")}<br/>
        <b>Visualization:</b> {config.get("visualization", "table")}<br/>
        """

        if "filters" in config and config["filters"]:
            config_text += f"<b>Filters:</b> {str(config['filters'])}<br/>"

        story.append(Paragraph(config_text, styles["Normal"]))

        # Build PDF
        doc.build(story)
        buffer.seek(0)

        return buffer.getvalue()
